#!/usr/bin/env python3
"""
功能清单 MD -> Excel 导出工具

解析功能清单 Markdown 文件中的表格，生成格式化的 Excel 文件。
支持一级模块合并行显示、表头加粗、自动列宽。

用法：
  python scripts/export_feature_list.py <path-to-feature-list.md>
  python scripts/export_feature_list.py <path-to-feature-list.md> --out <output.xlsx>

依赖：pip install openpyxl
"""

import re
import sys
import argparse
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Run: pip install openpyxl")
    sys.exit(1)


# ── 配色常量 ────────────────────────────────────────────
HEADER_BG = "2563EB"   # 蓝色表头背景
HEADER_FG = "FFFFFF"   # 白色表头字体
MODULE_BG = "EFF6FF"   # 一级模块行浅蓝背景
P0_COLOR  = "DC2626"   # P0 红色
P1_COLOR  = "D97706"   # P1 橙色
P2_COLOR  = "16A34A"   # P2 绿色

THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def parse_md_table(lines: list[str]) -> list[list[str]]:
    """解析 Markdown 表格行，返回单元格列表。"""
    rows = []
    for line in lines:
        line = line.strip()
        if not line.startswith("|"):
            continue
        cells = [c.strip() for c in line.split("|")[1:-1]]
        if cells and not all(re.match(r"^[-:\s]+$", c) for c in cells):
            rows.append(cells)
    return rows


def find_feature_table(content: str) -> list[list[str]] | None:
    """在 MD 中定位功能清单表格（含「二级功能」或「功能点」列的表格）。"""
    lines = content.splitlines()
    in_table = False
    table_lines: list[str] = []

    for line in lines:
        stripped = line.strip()
        if not stripped.startswith("|"):
            if in_table and table_lines:
                break
            continue
        cells = [c.strip() for c in stripped.split("|")[1:-1]]
        if not cells:
            continue
        header_text = str(cells)
        if "二级功能" in header_text or "功能点" in header_text or "功能模块" in header_text:
            in_table = True
            table_lines = [line]
            continue
        if in_table:
            table_lines.append(line)

    if not table_lines:
        return None
    return parse_md_table(table_lines)


def find_meta_table(content: str) -> dict[str, str]:
    """从「基本信息」表格中提取元数据。"""
    meta: dict[str, str] = {}
    lines = content.splitlines()
    in_meta = False
    for line in lines:
        stripped = line.strip()
        if "基本信息" in stripped:
            in_meta = True
            continue
        if in_meta:
            if not stripped.startswith("|"):
                break
            cells = [c.strip() for c in stripped.split("|")[1:-1]]
            if len(cells) >= 2 and not re.match(r"^[-:\s]+$", cells[0]):
                meta[cells[0]] = cells[1]
    return meta


def detect_columns(header_row: list[str]) -> dict[str, int]:
    """检测各列的位置（0-based）。"""
    col_map: dict[str, int] = {}
    for i, h in enumerate(header_row):
        h_lower = h.lower()
        if "序号" in h or "编号" in h:
            col_map["id"] = i
        elif "一级" in h or "功能模块" in h:
            col_map["module"] = i
        elif "二级" in h or "功能点" in h:
            col_map["feature"] = i
        elif "描述" in h:
            col_map["desc"] = i
        elif "优先" in h:
            col_map["priority"] = i
        elif "来源" in h:
            col_map["source"] = i
        elif "备注" in h or "remark" in h_lower:
            col_map["remark"] = i
    return col_map


def build_excel(rows: list[list[str]], meta: dict[str, str], out_path: Path) -> None:
    """将解析后的行数据写入格式化 Excel。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "功能清单"

    if not rows:
        wb.save(out_path)
        return

    header_row = rows[0]
    col_map = detect_columns(header_row)
    module_col_idx = col_map.get("module", -1)

    # ── 写入表头 ─────────────────────────────────────────
    header_font = Font(name="微软雅黑", bold=True, color=HEADER_FG, size=11)
    header_fill = PatternFill(fill_type="solid", fgColor=HEADER_BG)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, cell_val in enumerate(header_row, 1):
        cell = ws.cell(row=1, column=col_idx, value=cell_val)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 24

    # ── 写入数据行 ───────────────────────────────────────
    module_fill = PatternFill(fill_type="solid", fgColor=MODULE_BG)
    normal_font = Font(name="微软雅黑", size=10)
    module_font = Font(name="微软雅黑", size=10, bold=True)
    wrap_align = Alignment(vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="top")

    current_module = ""
    module_start_row = 2
    module_rows: list[tuple[int, int]] = []   # (start, end) 待合并区间

    data_rows = rows[1:]
    for row_idx, row in enumerate(data_rows, 2):
        # 补齐列数
        while len(row) < len(header_row):
            row.append("")

        # 检查是否一级模块行（二级功能为空）
        is_module_row = False
        if module_col_idx >= 0 and len(row) > module_col_idx:
            module_val = row[module_col_idx]
            feature_val = row[col_map.get("feature", module_col_idx + 1)] if col_map.get("feature") is not None else ""
            if module_val and not feature_val:
                is_module_row = True

        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER

            if is_module_row:
                cell.font = module_font
                cell.fill = module_fill
                cell.alignment = center_align if col_idx == 1 else Alignment(vertical="top", wrap_text=True)
            else:
                cell.font = normal_font
                # 优先级着色
                actual_col = col_idx - 1
                if actual_col == col_map.get("priority"):
                    if val == "P0":
                        cell.font = Font(name="微软雅黑", size=10, bold=True, color=P0_COLOR)
                    elif val == "P1":
                        cell.font = Font(name="微软雅黑", size=10, bold=True, color=P1_COLOR)
                    elif val == "P2":
                        cell.font = Font(name="微软雅黑", size=10, bold=True, color=P2_COLOR)
                    cell.alignment = center_align
                else:
                    cell.alignment = wrap_align

        # 一级模块合并追踪
        if module_col_idx >= 0 and row_idx > 2:
            row_module = row[module_col_idx] if len(row) > module_col_idx else ""
            if row_module and row_module != current_module:
                if current_module and not is_module_row:
                    module_rows.append((module_start_row, row_idx - 1))
                current_module = row_module
                module_start_row = row_idx

    # 最后一组
    if current_module and module_col_idx >= 0:
        module_rows.append((module_start_row, len(data_rows) + 1))

    # 合并一级模块列（相同模块跨多行时）
    for start, end in module_rows:
        if end > start and module_col_idx >= 0:
            try:
                merge_col = get_column_letter(module_col_idx + 1)
                ws.merge_cells(f"{merge_col}{start}:{merge_col}{end}")
                merged_cell = ws[f"{merge_col}{start}"]
                merged_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            except Exception:
                pass

    # ── 自动列宽 ─────────────────────────────────────────
    col_widths = {i: len(str(v)) for i, v in enumerate(header_row, 1)}
    for row in data_rows:
        for i, val in enumerate(row, 1):
            # 多行内容按最长行计算
            max_line_len = max((len(line) for line in str(val).splitlines()), default=0)
            col_widths[i] = max(col_widths.get(i, 0), max_line_len)

    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        # 描述列最宽 60，其他列最宽 35
        col_name = header_row[col_idx - 1] if col_idx <= len(header_row) else ""
        max_w = 60 if "描述" in col_name else 35
        ws.column_dimensions[col_letter].width = min(max(width + 2, 8), max_w)

    # ── 冻结首行 ─────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── 元数据 sheet ────────────────────────────────────
    if meta:
        ws_meta = wb.create_sheet(title="基本信息")
        ws_meta.column_dimensions["A"].width = 18
        ws_meta.column_dimensions["B"].width = 40
        for r, (k, v) in enumerate(meta.items(), 1):
            ws_meta.cell(row=r, column=1, value=k).font = Font(name="微软雅黑", bold=True, size=10)
            ws_meta.cell(row=r, column=2, value=v).font = Font(name="微软雅黑", size=10)

    wb.save(out_path)


def export(md_path: str | Path, out_path: str | Path | None = None) -> Path:
    md_path = Path(md_path)
    if not md_path.exists():
        raise FileNotFoundError(f"文件不存在: {md_path}")

    content = md_path.read_text(encoding="utf-8")
    rows = find_feature_table(content)
    if not rows:
        raise ValueError("未找到功能清单表格（需包含「二级功能」、「功能点」或「功能模块」列）")

    meta = find_meta_table(content)

    if out_path is None:
        out_path = md_path.with_suffix(".xlsx")
    else:
        out_path = Path(out_path)

    build_excel(rows, meta, out_path)
    return out_path


def main():
    parser = argparse.ArgumentParser(description="功能清单 MD -> Excel 导出工具")
    parser.add_argument("md_path", help="功能清单 Markdown 文件路径")
    parser.add_argument("--out", help="输出 Excel 文件路径（默认与 MD 文件同目录）")
    args = parser.parse_args()

    try:
        out = export(args.md_path, args.out)
        print(f"导出成功: {out}")
    except (FileNotFoundError, ValueError) as e:
        print(f"Error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
